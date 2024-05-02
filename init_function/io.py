# -*- coding: utf-8 -*-
## custom function for data I/O
from configparser import ConfigParser
import openpyxl

def Read_config(config_file):
    con=ConfigParser()
    config=con.read('test.ini', encoding='UTF-8')
    return config

def Readxlsx(xlsx_path):
    workbook=openpyxl.load_workbook(xlsx_path)
    sheet= workbook.worksheets[0]
    maxcol=sheet.max_column
    mincol=sheet.min_column 
    minrow=sheet.min_row 
    maxrow=sheet.max_row
    data=sheet.iter_cols(min_row=1,values_only=True)
    return(data)

def Getdata(data,type='xy'):
    xdata=[]
    ydata=[]
    xlabel=""
    ylabel=""
    xydata=[]
    for data_val in data:
        if(data_val[0]=="x"):
            xdata=data_val
            xdata=list(xdata)
            xlabel=xdata[1]
            del xdata[0]
            del xdata[0]
            xdata=[x for x in xdata if x != None]
            continue
        elif(data_val[0]=='y'):
            ydata=data_val
            ydata=list(ydata)
            ylabel=ydata[1]
            del ydata[0]
            del ydata[0]
            ydata=[y for y in ydata if y !=None]
            if type=='xy':
                xydata.append([xdata,ydata,xlabel,ylabel])
    return(xydata)
