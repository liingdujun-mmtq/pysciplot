# -*- coding: utf-8 -*-
## custom function for data I/O
from configparser import ConfigParser
import openpyxl

## read config file or preset file, and return the selected section in config file
def Read_config(config_file,section):
    preset=ConfigParser()
    preset.read(config_file, encoding='UTF-8')
    config=ConfigParser()
    config.read("config.conf", encoding='UTF-8')
    if config.has_section(section):
        loadconfig=config[section]
    else:
        loadconfig=None
    loadpreset=preset[section]
    return [loadconfig,loadpreset]

## read  selected section from config and preset file
# return config setting, if config= None, return preset
# if setting is a list, return the selected item (via num)
def Get_item(section,item,num=0):
    ## split: split str to list if "," in str
    def split(str,num=0):
        if ',' in str:
            all_str=str.split(',')
            max_num=len(all_str)
            return all_str[num%max_num]
        else:
            return str
    if section[0]==None:
        preset_item=section[1].get(item)
        return split(preset_item,num)
    else:
        config_item=section[0].get(item)
        preset_item=section[1].get(item)
        if config_item != None:
            return split(config_item,num)
        else:
            return split(preset_item,num)

## Class used to read config file
class preset(object):
    def __init__(self, section):
        self.section = section

class preset_plot_plot(preset):
    def __init__(self, section,num=0):
        self.linestyle=Get_item(section,"linestyle",num)
        self.color=Get_item(section,"color",num)
        self.linewidth=Get_item(section,"linewidth",num)
        self.marker=Get_item(section,'marker',num)
        self.markersize=Get_item(section,"markersize",num)
        self.label_fontsize=Get_item(section,'label_fontsize',num)

class preset_legend(preset):
    def __init__(self, section):
        self.legend_location=Get_item(section,"legend_location")
        self.fontsize=Get_item(section,"fontsize")
        self.edgecolor=Get_item(section,"edgecolor")     


## Custom class used in configure
class configure(preset):
    def __init__(self, section):
        self.xlabel=Get_item(section,"xlabel")
        self.ylabel=Get_item(section,"ylabel")

## Read data from xlsx file
def Readxlsx(xlsx_path):
    workbook=openpyxl.load_workbook(xlsx_path,data_only=True)
    sheet= workbook.worksheets[0]
    maxcol=sheet.max_column
    mincol=sheet.min_column 
    minrow=sheet.min_row 
    maxrow=sheet.max_row
    data=sheet.iter_cols(min_row=1,values_only=True)
    return(data)

def Getdata(data,type='xy'):
    xdata=[]
    ydata=[]
    xlabel=""
    ylabel=""
    xydata=[]
    for data_val in data:
        if(data_val[0]=="x"):
            xdata=data_val
            xdata=list(xdata)
            xlabel=xdata[1]
            del xdata[0]
            del xdata[0]
            xdata=[x for x in xdata if x != None]
            continue
        elif(data_val[0]=='y'):
            ydata=data_val
            ydata=list(ydata)
            ylabel=ydata[1]
            del ydata[0]
            del ydata[0]
            ydata=[y for y in ydata if y !=None]
            if type=='xy':
                xydata.append([xdata,ydata,ylabel])
    return(xydata)
